from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
#wb = load_workbook('grades.xlsx')

#ws = wb.active
#print(wb.sheetnames)
#ws['A2'].value = "Test" #Name is saved after the workbook is saved
#wb.save('grades.xlsx')

#wb.create_sheet("Test")
#print(wb.sheetnames)

# wb = Workbook()
# ws = wb.active
# ws.title = "Data"
# ws.append(['Today', 'is', 'Wednesday', '!'])
# ws.append(['Tomorrow', 'is', 'Tuesday', '!'])
# wb.save('wednesday.xlsx')

# wb = load_workbook('wednesday.xlsx')
# ws = wb.active
#
# for row in range(1, 11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         ws[char + str(row)] = char + str(row)
#
# wb.save('wednesday.xlx')


# merging cells

wb = load_workbook('grades.xlsx')
ws = wb.active

#ws.merge_cells("A1:D1")  #gets rid of data in cells B1 and so on till D1

#ws.delete_rows(1)

#ws.insert_cols(2) #insers column B
#ws.delete_cols(2) #delete column b

ws.move_range("C1:D11", rows=2, cols=2)


wb.save('grades.xlsx')